"""Convert the audited 保健 workbook into data/health-items.json."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook

SHEET_NAME = "問題集"
HEADERS = ("No.", "問題", "答え", "重要度", "出典", "出題タイプ", "図表・資料")

# The workbook records the textbook page each question comes from; the app
# groups questions by the page spreads it filters and reports on.
PAGE_RANGES = {
    "12": "p.12–13", "13": "p.12–13",
    "14": "p.14–15", "15": "p.14–15",
    "16": "p.16–17", "17": "p.16–17",
    "20": "p.20–21", "21": "p.20–21",
    "24": "p.24–25", "25": "p.24–25",
    "26": "p.26–27", "27": "p.26–27",
    "30": "p.30–31", "31": "p.30–31",
    "34": "p.34–35", "35": "p.34–35",
}

RANGE_ORDER = [
    "p.12–13", "p.14–15", "p.16–17", "p.20–21",
    "p.24–25", "p.26–27", "p.30–31", "p.34–35",
]

KINDS = {"語句": "用語", "図表": "図表", "資料": "資料"}

SOURCE_PATTERN = re.compile(r"^(写真\d+)・p\.(\d+)(.+)$")


def clean_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def parse_source(label: str) -> tuple[str, str, str]:
    """Split 写真1・p.12図2 into (range, printable source, photo label)."""
    match = SOURCE_PATTERN.match(clean_text(label))
    if not match:
        raise ValueError(f"unreadable source label: {label!r}")
    photo, page, section = match.groups()
    if page not in PAGE_RANGES:
        raise ValueError(f"unknown page in source label: {label!r}")
    return PAGE_RANGES[page], f"p.{page} {section}", photo


def classify_kind(answer: str, question_type: str) -> str:
    if question_type in KINDS:
        return KINDS[question_type]
    return "年号" if re.fullmatch(r"\d{3,4}年", answer) else "数値"


def convert(workbook_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"{workbook_path}: sheet {SHEET_NAME} is required")
    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    header = tuple(clean_text(cell) for cell in next(rows)[: len(HEADERS)])
    if header != HEADERS:
        raise SystemExit(f"{workbook_path}: unexpected header {header}")

    items: list[dict[str, object]] = []
    seen: set[tuple[str, str]] = set()
    for row in rows:
        question = clean_text(row[1])
        answer = clean_text(row[2])
        if not question or not answer:
            continue
        importance = clean_text(row[3]) or "C"
        study_range, source, photo = parse_source(row[4])
        key = (question, answer)
        if key in seen:
            continue
        seen.add(key)

        kind = classify_kind(answer, clean_text(row[5]))
        items.append({
            "id": "",
            "subject": "health",
            "number": 0,
            "importance": importance,
            "english": question,
            "japanese": answer,
            "healthQuestion": question,
            "healthAnswer": answer,
            "type": "health-term",
            "answerFormat": "term",
            "kind": kind,
            "range": study_range,
            "lesson": study_range,
            "title": source,
            "source": source,
            "sourceDetail": source,
            "sources": [{"lesson": study_range, "title": source, "detail": photo}],
            "tags": [kind, "語句"],
            "acceptedAnswers": [answer],
            "questionModes": ["health_recall"],
        })

    items.sort(key=lambda item: RANGE_ORDER.index(item["range"]))
    for number, item in enumerate(items, start=1):
        item["number"] = number
        item["id"] = f"health-{number:04d}"
    return items


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    items = convert(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(items, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"Wrote {len(items)} items to {args.output}")


if __name__ == "__main__":
    main()
