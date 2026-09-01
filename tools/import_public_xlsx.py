"""Convert the audited 公共 workbook into data/public-items.json."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook

SHEET_NAME = "全問題"
HEADERS = ("No.", "問題", "答え", "重要度", "出典", "形式")

# The workbook labels each source by textbook photo; the app groups questions by
# the page spreads those photos cover.
PHOTO_RANGES = {
    "写真1": "p.36–37",
    "写真2": "p.40–47",
    "写真3": "p.40–47",
    "写真4": "p.40–47",
    "写真5": "p.40–47",
    "写真6": "p.60–63",
    "写真7": "p.60–63",
    "写真8": "p.68–69",
    "写真9": "p.70–73",
    "写真10": "p.70–73",
    "写真11": "p.76–77",
}

RANGE_ORDER = ["p.36–37", "p.40–47", "p.60–63", "p.68–69", "p.70–73", "p.76–77"]


def clean_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def parse_source(label: str) -> tuple[str, str, str]:
    """Split 写真1・p.36・本文中段 into (range, printable source, photo label)."""
    parts = [part for part in clean_text(label).split("・") if part]
    if not parts:
        raise ValueError(f"empty source label: {label!r}")
    photo = parts[0]
    if photo not in PHOTO_RANGES:
        raise ValueError(f"unknown photo label: {label!r}")
    printable = "・".join(parts[1:]) or photo
    return PHOTO_RANGES[photo], printable, photo


INSTITUTION_ANSWERS = {
    "国会", "内閣", "裁判所", "衆議院", "参議院", "最高裁判所", "地方裁判所",
    "高等裁判所", "地方議会", "内閣総理大臣",
}


def classify_kind(question: str, answer: str, source: str) -> str:
    if question.endswith("は誰か。"):
        return "人物"
    if re.fullmatch(r"[『「].+[』」]", answer):
        return "作品"
    if re.fullmatch(r"\d{3,4}年(\d{1,2}月)?(\d{1,2}日)?", answer):
        return "年号"
    if re.fullmatch(r"[\d.,]+(%|％|人|年|倍|割|議席|票)?", answer):
        return "数値"
    if re.search(r"(順位|ランキング|何位)", question):
        return "順位"
    if re.search(r"(条約|規約|宣言|憲章|議定書)$", answer):
        return "条約"
    if answer in INSTITUTION_ANSWERS or re.search(r"(機関|議院|裁判所|委員会|省|庁|会議)$", answer):
        return "機関"
    if re.search(r"(都|道|府|県|市|町|村)$", answer):
        return "地名"
    if len(answer) >= 3 and re.search(r"(法|条例)$", answer) and "性質" not in question:
        return "法律"
    if re.search(r"(原理|原則)$", answer) or re.search(r"(原理|原則)を何というか", question):
        return "原理"
    if re.search(r"(権|自由)$", answer) and re.search(r"(権利|自由)", question):
        return "人権"
    if re.search(r"(制度|しくみ|方式|制)を何というか", question):
        return "制度"
    if re.search(r"(図\d|表\d|グラフ|年表)", source):
        return "図表"
    if "トピック" in source:
        return "資料"
    return "用語"


def convert(workbook_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"{workbook_path}: sheet {SHEET_NAME} is required")
    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    header = tuple(clean_text(cell) for cell in next(rows)[: len(HEADERS)])
    if header != HEADERS:
        raise SystemExit(f"{workbook_path}: unexpected header {header}")

    items: list[dict[str, object]] = []
    seen: set[tuple[str, str]] = set()
    for row in rows:
        question = clean_text(row[1])
        answer = clean_text(row[2])
        if not question or not answer:
            continue
        importance = clean_text(row[3]) or "C"
        study_range, source, photo = parse_source(row[4])
        key = (question, answer)
        if key in seen:
            continue
        seen.add(key)

        number = len(items) + 1
        kind = classify_kind(question, answer, clean_text(row[4]))
        items.append({
            "id": f"public-{number:04d}",
            "subject": "public",
            "number": number,
            "importance": importance,
            "english": question,
            "japanese": answer,
            "publicQuestion": question,
            "publicAnswer": answer,
            "type": "public-term",
            "answerFormat": "term",
            "kind": kind,
            "range": study_range,
            "lesson": study_range,
            "title": source,
            "source": source,
            "sourceDetail": source,
            "sources": [{"lesson": study_range, "title": source, "detail": photo}],
            "tags": [kind, "語句"],
            "acceptedAnswers": [answer],
            "questionModes": ["public_recall"],
        })

    items.sort(key=lambda item: (RANGE_ORDER.index(item["range"]), item["number"]))
    for number, item in enumerate(items, start=1):
        item["number"] = number
        item["id"] = f"public-{number:04d}"
    return items


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    items = convert(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(items, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"Wrote {len(items)} items to {args.output}")


if __name__ == "__main__":
    main()
