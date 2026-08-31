#!/usr/bin/env python3
"""Convert the supplied vocabulary workbook into the app's JSON data model."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


SOURCE_MAP = {
    "OriHime": {
        "range": "OriHime",
        "slug": "orihime",
        "lesson": "OriHime",
        "title": "A Vehicle of Your Heart",
    },
    "Mars": {
        "range": "Mars",
        "slug": "mars",
        "lesson": "Mars",
        "title": "Human Habitation on Mars",
    },
    "Kakigori": {
        "range": "Kakigori",
        "slug": "kakigori",
        "lesson": "Kakigori",
        "title": "A Cool Food: Kakigori",
    },
    "Snow": {
        "range": "Snow",
        "slug": "snow",
        "lesson": "Lesson 4",
        "title": "Snow",
    },
    "Lesson 2": {
        "range": "Plastic",
        "slug": "plastic",
        "lesson": "Lesson 2",
        "title": "Plastic",
    },
    "Lesson 3": {
        "range": "FOMO",
        "slug": "fomo",
        "lesson": "Lesson 3",
        "title": "FOMO",
    },
    "Lesson 4": {
        "range": "Snow",
        "slug": "snow",
        "lesson": "Lesson 4",
        "title": "Snow",
    },
    "Lesson 5": {
        "range": "Shinkansen",
        "slug": "shinkansen",
        "lesson": "Lesson 5",
        "title": "Shinkansen",
    },
    "Lesson 6": {
        "range": "Taste Buds",
        "slug": "taste-buds",
        "lesson": "Lesson 6",
        "title": "Taste Buds",
    },
}

SOURCE_RE = re.compile(
    r"^(OriHime|Mars|Kakigori|Snow|Lesson [2-6])"
    r"(?:\s+p\.([^\s]+))?(?:\s+(?:¶(.+)|(.+)))?$"
)
NEW_SOURCE_RE = re.compile(
    r"^Scan p\.(\d+) / (.+) / (textbook|book) pp\.([0-9]+(?:[–-][0-9]+)?)$"
)
NEW_SOURCE_MAP = {
    "OriHime — A Vehicle of Your Heart": {
        "range": "OriHime",
        "slug": "orihime",
        "lesson": "OriHime",
        "title": "A Vehicle of Your Heart",
    },
    "Human Habitation on Mars": {
        "range": "Mars",
        "slug": "mars",
        "lesson": "Mars",
        "title": "Human Habitation on Mars",
    },
    "A Cool Food: Kakigori": {
        "range": "Kakigori",
        "slug": "kakigori",
        "lesson": "Kakigori",
        "title": "A Cool Food: Kakigori",
    },
    "Lesson 2: Plastic Pollution": {
        "range": "Plastic",
        "slug": "plastic",
        "lesson": "Lesson 2",
        "title": "Plastic Pollution",
    },
    "Lesson 3: FOMO": {
        "range": "FOMO",
        "slug": "fomo",
        "lesson": "Lesson 3",
        "title": "FOMO",
    },
    "Lesson 4: Snow": {
        "range": "Snow",
        "slug": "snow",
        "lesson": "Lesson 4",
        "title": "Snow",
    },
    "Lesson 5: 7-Minute Miracle": {
        "range": "Shinkansen",
        "slug": "shinkansen",
        "lesson": "Lesson 5",
        "title": "7-Minute Miracle",
    },
    "Lesson 6: Taste & Supertasters": {
        "range": "Taste Buds",
        "slug": "taste-buds",
        "lesson": "Lesson 6",
        "title": "Taste & Supertasters",
    },
}
AUDIT_SOURCE_RE = re.compile(r"^P([1-8])(?:-TITLE|-¶(\d+)-S(\d+))$")
AUDIT_SOURCE_MAP = {
    "1": {"range": "OriHime", "slug": "orihime", "lesson": "OriHime", "title": "A Vehicle of Your Heart", "page": "62–63"},
    "2": {"range": "Mars", "slug": "mars", "lesson": "Mars", "title": "Human Habitation on Mars", "page": "106–107"},
    "3": {"range": "Kakigori", "slug": "kakigori", "lesson": "Kakigori", "title": "A Cool Food: Kakigori", "page": "46–47"},
    "4": {"range": "Plastic", "slug": "plastic", "lesson": "Lesson 2", "title": "Plastic Garbage", "page": "4–5"},
    "5": {"range": "FOMO", "slug": "fomo", "lesson": "Lesson 3", "title": "FOMO", "page": "6–7"},
    "6": {"range": "Snow", "slug": "snow", "lesson": "Lesson 4", "title": "Why Snow Is White", "page": "8–9"},
    "7": {"range": "Shinkansen", "slug": "shinkansen", "lesson": "Lesson 5", "title": "7-Minute Miracle", "page": "10–11"},
    "8": {"range": "Taste Buds", "slug": "taste-buds", "lesson": "Lesson 6", "title": "Taste Buds", "page": "12–13"},
}
IMPORTANCE_ORDER = ["SSS", "SS", "S", "A", "B", "C"]
DIFFICULTY_ORDER = ["A", "B", "C", "D", "E", "F"]
APP_WORD_IMPORTANCE = {"SSS", "SS", "S", "A", "B"}
WORD_RE = re.compile(r"[A-Za-z]+(?:['’\-][A-Za-z]+)*")
PAREN_RE = re.compile(r"[（(]([^（()）]*)[)）]")
JAPANESE_RE = re.compile(r"[ぁ-んァ-ヶ一-龥]")
NAME_NOTE_RE = re.compile(r"^[^、]*(?:名|の一部)$")
PLACEHOLDERS = {"a", "b", "s", "v"}
PREPOSITIONS = {
    "about",
    "above",
    "across",
    "after",
    "against",
    "along",
    "among",
    "around",
    "as",
    "at",
    "before",
    "below",
    "beneath",
    "beside",
    "between",
    "beyond",
    "by",
    "despite",
    "during",
    "except",
    "for",
    "from",
    "in",
    "inside",
    "into",
    "like",
    "near",
    "of",
    "on",
    "onto",
    "outside",
    "over",
    "past",
    "since",
    "through",
    "throughout",
    "to",
    "toward",
    "towards",
    "under",
    "underneath",
    "until",
    "upon",
    "with",
    "within",
    "without",
}


def clean_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def parse_source(label: str) -> dict[str, str]:
    cleaned = clean_text(label)
    audit_match = AUDIT_SOURCE_RE.match(cleaned)
    if audit_match:
        source_key, paragraph, sentence = audit_match.groups()
        source = AUDIT_SOURCE_MAP[source_key]
        location = "title" if "TITLE" in cleaned else f"¶{paragraph} S{sentence}"
        return {
            "range": source["range"],
            "lesson": source["lesson"],
            "title": source["title"],
            "page": source["page"],
            "paragraph": paragraph or "",
            "note": "TITLE" if "TITLE" in cleaned else f"S{sentence}",
            "detail": f"textbook pp.{source['page']} · {location}",
            "label": cleaned,
            "slug": source["slug"],
        }

    new_match = NEW_SOURCE_RE.match(cleaned)
    if new_match:
        scan_page, source_key, book_kind, page = new_match.groups()
        if source_key not in NEW_SOURCE_MAP:
            raise ValueError(f"Unrecognized source title: {source_key!r}")
        source = NEW_SOURCE_MAP[source_key]
        return {
            "range": source["range"],
            "lesson": source["lesson"],
            "title": source["title"],
            "page": page,
            "paragraph": "",
            "note": f"Scan p.{scan_page}",
            "detail": f"Scan p.{scan_page} / {book_kind} pp.{page}",
            "label": cleaned,
            "slug": source["slug"],
        }

    match = SOURCE_RE.match(cleaned)
    if not match:
        raise ValueError(f"Unrecognized source label: {label!r}")
    source_key, page, paragraph, note = match.groups()
    source = SOURCE_MAP[source_key]
    detail_parts = []
    if page:
        detail_parts.append(f"p.{page}")
    if paragraph:
        detail_parts.append(f"¶{paragraph}")
    if note:
        detail_parts.append(note)
    return {
        "range": source["range"],
        "lesson": source["lesson"],
        "title": source["title"],
        "page": page or "",
        "paragraph": paragraph or "",
        "note": note or "",
        "detail": " ".join(detail_parts),
        "label": clean_text(label),
        "slug": source["slug"],
    }


def accepted_answers(english: str) -> list[str]:
    answers = [clean_text(english)]
    if re.search(r"\s+/\s+", english):
        answers = [clean_text(part) for part in re.split(r"\s+/\s+", english)]

    variants: list[str] = []
    for answer in answers:
        variants.append(answer)
        without_ellipsis = re.sub(r"(?:\s*\.{3}|\s*…+|\s*～+)$", "", answer).strip()
        if without_ellipsis and without_ellipsis != answer:
            variants.append(without_ellipsis)

    unique: list[str] = []
    seen: set[str] = set()
    for answer in variants:
        key = answer.casefold()
        if answer and key not in seen:
            seen.add(key)
            unique.append(answer)
    return unique


def primary_notation(english: str) -> str:
    """「help A (to) do / help A + 原形」のような併記は最初の表記を代表にする。"""
    return clean_text(re.split(r"\s+/\s+", english)[0])


def usage_key(english: str) -> str:
    """任意扱いの括弧と併記を落とし、同じ語法の書き分けを1件にまとめるためのキー。

    「help A (to) do / help A + 原形」「help A (to) do」「help A do」はすべて
    「help a do」になり、訳も出典も1件に統合される。
    """
    stripped = PAREN_RE.sub(" ", primary_notation(english))
    return re.sub(r"\s+", " ", stripped).strip().casefold()


def content_tokens(text: str) -> set[str]:
    """語法のA・B・S・Vのような1文字のプレースホルダを除いた英語トークン。"""
    return {
        match.group().casefold()
        for match in word_matches(text)
        if len(match.group()) > 1
    }


def strip_answer_notes(japanese: str, english: str) -> str:
    """訳の中に答えがまるごと出ている注記を取り除く。

    「NASA（米国航空宇宙局）」のように答えの後ろへ訳が続く形は括弧の中を訳として採り、
    「味蕾（taste bud）」のような補足注記や「bullet trainで新幹線」のような用例は落とす。
    答えの一部しか出ていない「NASAのコンテストで優勝した」のような訳はそのまま残す。
    """
    answer = content_tokens(english)
    if not answer:
        return japanese

    def leaks(text: str) -> bool:
        return answer <= content_tokens(text)

    variants: list[str] = []
    for variant in japanese.split("／"):
        variant = PAREN_RE.sub(lambda m: "" if leaks(m.group(1)) else m.group(), variant)
        if PAREN_RE.search(variant) and leaks(PAREN_RE.sub("", variant)):
            notes = [note for note in PAREN_RE.findall(variant) if note and not leaks(note)]
            variant = "、".join(note for note in notes if JAPANESE_RE.search(note))
        cleaned = "、".join(
            segment for segment in variant.split("、") if segment and not leaks(segment)
        ).strip("、 ")
        if cleaned:
            variants.append(cleaned)
    return "／".join(dict.fromkeys(variants))


def is_quizzable(japanese: str, original: str) -> bool:
    """訳が残っていて、かつ「人名」「〜の一部」のような名前ラベルだけでないか。"""
    variants = [variant for variant in japanese.split("／") if variant]
    if not variants:
        return False
    labels = PAREN_RE.findall(original) + variants
    return not any(NAME_NOTE_RE.match(label) for label in labels)


def classify_type(source_type: str, english: str) -> str:
    if source_type == "単語":
        return "word"
    if source_type == "熟語":
        return "phrase"
    if source_type == "構文":
        return "structure"
    if source_type == "語法":
        return "structure"
    if re.search(r"\b(?:A|B|S|V)\b", english):
        return "structure"
    structural_patterns = (
        r"\bto do\b",
        r"(?:～|\.\.\.)ing\b",
        r"^There (?:is|are)\b",
        r"^It (?:is|was)\b.*\bthat\b",
        r"\bthe (?:comparative|more|less)\b",
    )
    if any(re.search(pattern, english, re.IGNORECASE) for pattern in structural_patterns):
        return "structure"
    return "expression"


def word_matches(text: str) -> list[re.Match[str]]:
    return list(WORD_RE.finditer(text))


def replace_span(text: str, start: int, end: int) -> str:
    return clean_text(f"{text[:start]} ___ {text[end:]}")


def build_preposition_blank(answer: str) -> dict[str, str] | None:
    matches = word_matches(answer)
    eligible: list[re.Match[str]] = []
    for index, match in enumerate(matches):
        token = match.group(0).casefold()
        if token not in PREPOSITIONS:
            continue
        if token == "to" and index + 1 < len(matches):
            next_token = matches[index + 1].group(0).casefold()
            if next_token in {"do", "doing"}:
                continue
        eligible.append(match)
    if not eligible:
        return None
    target = eligible[-1]
    return {
        "prompt": replace_span(answer, target.start(), target.end()),
        "answer": target.group(0),
    }


def build_phrase_blank(answer: str) -> dict[str, str] | None:
    matches = word_matches(answer)
    if len(matches) < 2:
        return None

    placeholder_indexes = {
        index
        for index, match in enumerate(matches)
        if match.group(0).casefold() in PLACEHOLDERS
        and len(match.group(0)) == 1
    }

    if placeholder_indexes:
        segments: list[list[int]] = []
        current: list[int] = []
        for index in range(len(matches)):
            if index in placeholder_indexes:
                if current:
                    segments.append(current)
                    current = []
            else:
                current.append(index)
        if current:
            segments.append(current)
        usable = [segment for segment in segments if segment and segment[0] > 0]
        selected = usable[-1] if usable else segments[-1]
        selected = selected[-2:]
    else:
        selected = list(range(max(1, len(matches) - 2), len(matches)))

    # 「help A (to) do」の「(to)」のような任意扱いの語は空欄に含めない。
    # 括弧を跨いで区切ると「help A ( ___」のような穴埋めになってしまう。
    brackets = [match.span() for match in PAREN_RE.finditer(answer)]
    selected = [
        index
        for index in selected
        if not any(
            start <= matches[index].start() and matches[index].end() <= end
            for start, end in brackets
        )
    ]
    if not selected:
        return None

    first = matches[selected[0]]
    last = matches[selected[-1]]
    blank_answer = clean_text(answer[first.start() : last.end()])
    if not blank_answer:
        return None
    return {
        "prompt": replace_span(answer, first.start(), last.end()),
        "answer": blank_answer,
    }


def best_importance(values: list[str]) -> str:
    return min(values, key=lambda value: IMPORTANCE_ORDER.index(value))


def hardest_difficulty(values: list[str]) -> str:
    known = [value for value in values if value in DIFFICULTY_ORDER]
    return max(known, key=lambda value: DIFFICULTY_ORDER.index(value)) if known else "—"


def parse_sources(value: object) -> list[dict[str, str]]:
    return [parse_source(part) for part in clean_text(value).split(";") if clean_text(part)]


def append_unique(target: list[str], values: list[str]) -> None:
    seen = {value.casefold() for value in target}
    for value in values:
        cleaned = clean_text(value)
        if cleaned and cleaned.casefold() not in seen:
            target.append(cleaned)
            seen.add(cleaned.casefold())


def append_unique_sources(
    target: list[dict[str, str]], values: list[dict[str, str]]
) -> None:
    seen = {source["label"] for source in target}
    for source in values:
        if source["label"] not in seen:
            target.append(source)
            seen.add(source["label"])


def quizzable_records(records: list[dict[str, object]]) -> list[dict[str, object]]:
    """答えが訳にそのまま出ている注記を落とし、出題として成立しない語句を除く。"""
    kept: list[dict[str, object]] = []
    for record in records:
        original = "／".join(record["japanese"])
        japanese = strip_answer_notes(original, str(record["english"]))
        if not is_quizzable(japanese, original):
            continue
        record["japanese"] = japanese.split("／")
        kept.append(record)
    return kept


def make_item(record: dict[str, object], order: int) -> dict[str, object]:
    english = str(record["english"])
    japanese = "／".join(record["japanese"])
    source_type = str(record["source_type"])
    sources = record["sources"]
    primary = sources[0]
    digest_source = f"{source_type}|{english}".casefold().encode("utf-8")
    digest = hashlib.sha1(digest_source).hexdigest()[:10]
    item_id = f"{primary['slug']}_{digest}"

    answers = accepted_answers(english)
    primary_answer = answers[0]
    item_type = classify_type(source_type, english)
    preposition_blank = build_preposition_blank(primary_answer)
    phrase_blank = (
        build_phrase_blank(primary_answer)
        if item_type in {"phrase", "structure", "expression"}
        else None
    )

    modes = ["en_to_ja_choice", "ja_to_en_choice", "ja_to_en_input"]
    tags = [item_type]
    if item_type == "word" and len(word_matches(primary_answer)) == 1:
        modes.append("spelling_input")
        tags.append("spelling")
    if preposition_blank:
        modes.append("preposition_input")
        tags.append("preposition")
    if phrase_blank:
        modes.append("phrase_blank_input")
        tags.append("blank")

    item: dict[str, object] = {
        "id": item_id,
        "english": english,
        "japanese": japanese,
        "type": item_type,
        "sourceType": source_type,
        "importance": record["importance"],
        "difficulty": record["difficulty"],
        "range": primary["range"],
        "lesson": primary["lesson"],
        "title": primary["title"],
        "source": "; ".join(source["label"] for source in sources),
        "sourceDetail": primary["detail"],
        "sources": [
            {key: value for key, value in source.items() if key != "slug"}
            for source in sources
        ],
        "tags": tags,
        "acceptedAnswers": answers,
        "questionModes": modes,
        "order": order,
    }
    if record.get("lemma"):
        item["lemma"] = record["lemma"]
    if record.get("surface_forms"):
        item["surfaceForms"] = record["surface_forms"]
    if record.get("examples"):
        item["examples"] = record["examples"]
    if record.get("notes"):
        item["note"] = "／".join(record["notes"])

    blanks: dict[str, dict[str, str]] = {}
    if preposition_blank:
        blanks["preposition"] = preposition_blank
    if phrase_blank:
        blanks["phrase"] = phrase_blank
    if blanks:
        item["blanks"] = blanks
    return item


def convert_audit(workbook) -> list[dict[str, object]]:
    grouped: OrderedDict[str, dict[str, object]] = OrderedDict()

    word_sheet = workbook["02_英単語一覧"]
    word_rows = word_sheet.iter_rows(values_only=True)
    word_headers = {clean_text(value): index for index, value in enumerate(next(word_rows))}
    for row in word_rows:
        lemma = clean_text(row[word_headers["原形"]])
        if not lemma:
            continue
        key = f"単語|{lemma.casefold()}"
        record = grouped.setdefault(key, {
            "english": lemma,
            "lemma": lemma,
            "japanese": [],
            "source_type": "単語",
            "importances": [],
            "difficulties": [],
            "sources": [],
            "surface_forms": [],
            "examples": [],
            "notes": [],
        })
        append_unique(record["japanese"], [row[word_headers["本文中の意味"]]])
        append_unique(record["surface_forms"], [row[word_headers["本文中の形"]]])
        append_unique_sources(record["sources"], parse_sources(row[word_headers["出典"]]))
        record["importances"].append(clean_text(row[word_headers["重要度"]]))
        record["difficulties"].append(clean_text(row[word_headers["難易度"]]))

    records: list[dict[str, object]] = []
    for record in grouped.values():
        record["importance"] = best_importance(record.pop("importances"))
        record["difficulty"] = hardest_difficulty(record.pop("difficulties"))
        if record["importance"] in APP_WORD_IMPORTANCE:
            records.append(record)

    for sheet_name, source_type, columns in (
        (
            "03_熟語・連語一覧",
            "熟語",
            {"english": "表現", "japanese": "日本語訳", "example": "本文中の意味・使い方", "importance": "重要度", "difficulty": "難易度"},
        ),
        (
            "04_語法一覧",
            "語法",
            {"english": "語法", "japanese": "意味", "example": "本文中の該当箇所", "importance": "重要度"},
        ),
    ):
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = {clean_text(value): index for index, value in enumerate(next(rows))}
        local: OrderedDict[str, dict[str, object]] = OrderedDict()
        for row in rows:
            english = clean_text(row[headers[columns["english"]]])
            if not english:
                continue
            key = usage_key(english)
            record = local.setdefault(key, {
                "english": primary_notation(english),
                "japanese": [],
                "source_type": source_type,
                "importances": [],
                "difficulties": [],
                "sources": [],
                "surface_forms": [],
                "examples": [],
                "notes": [],
            })
            append_unique(record["japanese"], [row[headers[columns["japanese"]]]])
            example = clean_text(row[headers[columns["example"]]])
            append_unique(record["examples"], [example])
            append_unique(record["notes"], [f"本文例: {example}"] if example else [])
            append_unique_sources(record["sources"], parse_sources(row[headers["出典"]]))
            record["importances"].append(clean_text(row[headers[columns["importance"]]]))
            if "difficulty" in columns:
                record["difficulties"].append(clean_text(row[headers[columns["difficulty"]]]))

        for record in local.values():
            record["importance"] = best_importance(record.pop("importances"))
            record["difficulty"] = hardest_difficulty(record.pop("difficulties"))
            records.append(record)

    return [
        make_item(record, order)
        for order, record in enumerate(quizzable_records(records), start=1)
    ]


def convert_legacy(workbook) -> list[dict[str, object]]:
    sheet = workbook["全一覧_重要度順"]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]
    modern_headers = {"重要度", "難易度", "種類", "英語", "日本語訳", "出典"}
    is_modern = modern_headers.issubset(headers)
    header_indexes = {header: index for index, header in enumerate(headers)}

    grouped: OrderedDict[str, dict[str, object]] = OrderedDict()

    for source_order, row in enumerate(rows, start=1):
        if not any(value is not None for value in row):
            continue

        if is_modern:
            importance_raw = row[header_indexes["重要度"]]
            difficulty_raw = row[header_indexes["難易度"]]
            source_type_raw = row[header_indexes["種類"]]
            english_raw = row[header_indexes["英語"]]
            japanese_raw = row[header_indexes["日本語訳"]]
            source_raw = row[header_indexes["出典"]]
            note_raw = row[header_indexes["備考"]] if "備考" in header_indexes else ""
        else:
            importance_raw, english_raw, japanese_raw, source_type_raw, source_raw = row[:5]
            difficulty_raw = ""
            note_raw = ""

        english = clean_text(english_raw)
        japanese = clean_text(japanese_raw)
        source_type = clean_text(source_type_raw)
        source_label = clean_text(source_raw)
        sources = [parse_source(part) for part in source_label.split(";")]
        key = english.casefold()

        if key not in grouped:
            grouped[key] = {
                "english": english,
                "japanese": [japanese],
                "source_type": source_type,
                "importance": clean_text(importance_raw),
                "difficulty": clean_text(difficulty_raw),
                "notes": [clean_text(note_raw)] if clean_text(note_raw) else [],
                "sources": sources,
                "source_order": source_order,
            }
            continue

        record = grouped[key]
        if japanese not in record["japanese"]:
            record["japanese"].append(japanese)
        existing_labels = {source["label"] for source in record["sources"]}
        record["sources"].extend(
            source for source in sources if source["label"] not in existing_labels
        )
        note = clean_text(note_raw)
        if note and note not in record["notes"]:
            record["notes"].append(note)

    items: list[dict[str, object]] = []
    for order, record in enumerate(quizzable_records(list(grouped.values())), start=1):
        english = record["english"]
        japanese = "／".join(record["japanese"])
        source_type = record["source_type"]
        sources = record["sources"]
        primary = sources[0]
        digest = hashlib.sha1(english.casefold().encode("utf-8")).hexdigest()[:10]
        item_id = f"{primary['slug']}_{digest}"

        answers = accepted_answers(english)
        primary_answer = answers[0]
        item_type = classify_type(source_type, english)
        preposition_blank = build_preposition_blank(primary_answer)
        phrase_blank = (
            build_phrase_blank(primary_answer)
            if item_type in {"phrase", "structure", "expression"}
            else None
        )

        modes = ["en_to_ja_choice", "ja_to_en_choice", "ja_to_en_input"]
        tags = [item_type]
        if item_type == "word" and len(word_matches(primary_answer)) == 1:
            modes.append("spelling_input")
            tags.append("spelling")
        if preposition_blank:
            modes.append("preposition_input")
            tags.append("preposition")
        if phrase_blank:
            modes.append("phrase_blank_input")
            tags.append("blank")

        item: dict[str, object] = {
            "id": item_id,
            "english": english,
            "japanese": japanese,
            "type": item_type,
            "sourceType": source_type,
            "importance": record["importance"],
            "difficulty": record["difficulty"],
            "range": primary["range"],
            "lesson": primary["lesson"],
            "title": primary["title"],
            "source": "; ".join(source["label"] for source in sources),
            "sourceDetail": primary["detail"],
            "sources": [
                {key: value for key, value in source.items() if key != "slug"}
                for source in sources
            ],
            "tags": tags,
            "acceptedAnswers": answers,
            "questionModes": modes,
            "order": order,
        }
        if record["notes"]:
            item["note"] = "／".join(record["notes"])
        blanks: dict[str, dict[str, str]] = {}
        if preposition_blank:
            blanks["preposition"] = preposition_blank
        if phrase_blank:
            blanks["phrase"] = phrase_blank
        if blanks:
            item["blanks"] = blanks
        items.append(item)

    return items


def convert(workbook_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if {"02_英単語一覧", "03_熟語・連語一覧", "04_語法一覧"}.issubset(workbook.sheetnames):
        return convert_audit(workbook)
    return convert_legacy(workbook)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    items = convert(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(items, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"Wrote {len(items)} items to {args.output}")


if __name__ == "__main__":
    main()
